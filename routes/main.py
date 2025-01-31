# routes/main.py
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
import tempfile
import os

from flask import Blueprint, render_template, redirect, url_for, request, jsonify
from flask_login import login_required, current_user
from config import CALCULATOR_CONFIG
from config import LANDING_PAGE_CONFIG
from calculations import calculate
from flask import send_file
import json

main = Blueprint("main", __name__)


@main.route("/")
def index():
    return render_template("landing.html", config=LANDING_PAGE_CONFIG)


@main.route("/profile")
@login_required
def profile():
    return render_template("profile.html")


@main.route("/calculator", methods=["GET", "POST"])
# @login_required
def calc():
    if request.method == "POST":
        results = calculate(request.form)
        import json

        raw_string = json.dumps(results, indent=4)
        return render_template("results.html", raw=raw_string, **results)
    return render_template("calculator.html", config=CALCULATOR_CONFIG)


@main.route("/imprint")
def imprint():
    return render_template("imprint.html", config=LANDING_PAGE_CONFIG)


@main.route("/privacy")
def privacy():
    return render_template("privacy.html", config=LANDING_PAGE_CONFIG)


@main.route("/download/pdf")
def download_pdf():
    data = request.args.get('data')
    if not data:
        return "No data provided", 400
    
    data = json.loads(data)
    
    # Create temporary file
    temp_dir = tempfile.gettempdir()
    pdf_path = os.path.join(temp_dir, "carbon_footprint.pdf")
    
    # Create PDF
    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph("Digital Carbon Footprint Report", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Main metrics
    main_data = [
        ["Metric", "Value (kg CO₂e)"],
        ["Data Center", f"{data['result_data_center']:.2f}"],
        ["Advertising", f"{data['result_advertising']:.2f}"],
        ["Content Delivery (Fixed Line)", f"{data['content_delivery_emissions']['fixed_line']:.2f}"],
        ["Content Delivery (Mobile)", f"{data['content_delivery_emissions']['mobile_line']:.2f}"]
    ]
    
    t = Table(main_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    return send_file(pdf_path, as_attachment=True, download_name="carbon_footprint.pdf")

@main.route("/download/excel")
def download_excel():
    data = request.args.get('data')
    if not data:
        return "No data provided", 400
    
    data = json.loads(data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Carbon Footprint"
    
    # Add headers
    ws['A1'] = "Category"
    ws['B1'] = "Subcategory"
    ws['C1'] = "Value"
    ws['D1'] = "Unit"
    
    # Add data
    row = 2
    
    # Main metrics
    ws[f'A{row}'] = "Data Center"
    ws[f'C{row}'] = data['result_data_center']
    ws[f'D{row}'] = "kg CO₂e"
    row += 1
    
    ws[f'A{row}'] = "Advertising"
    ws[f'C{row}'] = data['result_advertising']
    ws[f'D{row}'] = "kg CO₂e"
    row += 1
    
    # End User Devices
    for device in ['laptop', 'desktop', 'smartphone', 'tablet', 'tv', 'ereader']:
        ws[f'A{row}'] = "End User Devices"
        ws[f'B{row}'] = device.capitalize()
        ws[f'C{row}'] = data['end_user_devices'][f'{device}_emissions']
        ws[f'D{row}'] = "kg CO₂e"
        row += 1
        
    # Content Delivery
    ws[f'A{row}'] = "Content Delivery"
    ws[f'B{row}'] = "Fixed Line"
    ws[f'C{row}'] = data['content_delivery_emissions']['fixed_line']
    ws[f'D{row}'] = "kg CO₂e"
    row += 1
    
    ws[f'A{row}'] = "Content Delivery"
    ws[f'B{row}'] = "Mobile Line"
    ws[f'C{row}'] = data['content_delivery_emissions']['mobile_line']
    ws[f'D{row}'] = "kg CO₂e"
    
    # Style the worksheet
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 15
    
    # Save to temporary file
    temp_dir = tempfile.gettempdir()
    excel_path = os.path.join(temp_dir, "carbon_footprint.xlsx")
    wb.save(excel_path)
    
    return send_file(excel_path, as_attachment=True, download_name="carbon_footprint.xlsx")
